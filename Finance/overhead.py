import re
from datetime import date
from openpyxl import load_workbook, Workbook
from src import config

def main(filename):
    inputEvents = []
    currencyDict = {
        "RMB": "CNY",
        "CNY": "CNY",
        "GBP": "GBP",
        "EUR": "EUR",
        "USD": "USD"
    }
    defaultCurrency = "GBP"

    while True:
        priceInput = input('> ')

        if priceInput:
            search = re.search(r'\d[^.0-9]', priceInput)  # search for the position where the numeric value ends
            currencySearch = re.search(r'[A-Z]{3}',
                                       priceInput)  # search for three capital letters, (currency-like items) in the input, if not present use default currency.

            if currencySearch:
                for currency in currencyDict:
                    if currency in priceInput:
                        priceCurrency = currencyDict[currency]
                        break
            else:
                priceCurrency = defaultCurrency

            if search:
                endpos = search.span()
                price = priceInput[:endpos[0] + 1].strip()

            else:
                price = priceInput.strip()

            try:
                float(price)  # test if the price obtained is a proper number
            except:
                print("format incorrect, please enter again")
            else:
                cause = input("Cause:").strip()
                if re.search(r'!!',cause):
                    myCourse = True
                else:
                    myCourse = False

                logdate = date.today().strftime("%Y/%m/%d")
                inputEvents.append((price, priceCurrency, logdate, myCourse ,cause))
        else:
            break

    #move events stored in inputEvents to Excel Document
    try:
        workbook = load_workbook(filename = filename)
    except:
        workbook = Workbook()
        workbook['Sheet'].title = 'Overhead'
        sheet = workbook['Overhead']
        sheet['A1'] = 'price'
        sheet['B1'] = 'currency'
        sheet['C1'] = 'logdate'
        sheet['D1'] = 'own money?'
        sheet['E1'] = 'event'
        # Try-except handles the first time using the script, creating a new overhead

    sheet = workbook['Overhead']
    startline = (len(sheet['A']) + 1)
    for index, event in enumerate(inputEvents):
        currentRow = startline + index
        sheet['A' + str(currentRow)] = event[0]
        sheet['B' + str(currentRow)] = event[1]
        sheet['C' + str(currentRow)] = event[2]
        sheet['D' + str(currentRow)] = event[3]
        sheet['E' + str(currentRow)] = event[4]
    workbook.save(filename = filename)

if __name__ == '__main__':
    main(config.OVERHEAD_FILEPATH)